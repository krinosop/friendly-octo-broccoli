from django.contrib import admin
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from .models import Director, Actor, Play, Casting, Performance, ActorRole
from django.db.models import Count, Avg
from openpyxl import Workbook
from django.utils import timezone
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages

class TheaterAdminSite(admin.AdminSite):
    site_header = 'Администрирование театра'
    site_title = 'Театр'
    index_title = 'Управление театром'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('statistics/', self.admin_view(self.statistics_view), name='theater_statistics'),
        ]
        return custom_urls + urls

    def statistics_view(self, request):
        from .views import statistics_view
        return statistics_view(request)

    def each_context(self, request):
        context = super().each_context(request)
        context['statistics_url'] = request.build_absolute_uri('/admin/statistics/')
        return context

admin_site = TheaterAdminSite(name='admin')

@admin.register(Play, site=admin_site)
class PlayAdmin(admin.ModelAdmin):
    list_display = ('title', 'director', 'genre', 'duration')
    list_filter = ('genre', 'director')
    search_fields = ('title', 'description')
    actions = ['delete_selected']

@admin.register(Actor, site=admin_site)
class ActorAdmin(admin.ModelAdmin):
    list_display = ('first_name', 'last_name', 'gender', 'date_of_birth')
    list_filter = ('gender',)
    search_fields = ('first_name', 'last_name')
    actions = ['delete_selected']

@admin.register(Director, site=admin_site)
class DirectorAdmin(admin.ModelAdmin):
    list_display = ('first_name', 'last_name', 'gender', 'years_of_experience')
    list_filter = ('gender',)
    search_fields = ('first_name', 'last_name')
    actions = ['delete_selected']

@admin.register(Casting, site=admin_site)
class CastingAdmin(admin.ModelAdmin):
    list_display = ('actor', 'play', 'role', 'casting_date', 'status')
    list_filter = ('status', 'casting_date')
    search_fields = ('actor__first_name', 'actor__last_name', 'play__title')
    date_hierarchy = 'casting_date'
    actions = ['delete_selected']

@admin.register(Performance, site=admin_site)
class PerformanceAdmin(admin.ModelAdmin):
    list_display = ('play', 'date', 'status', 'tickets_available', 'ticket_price')
    list_filter = ('status', 'date', 'play')
    search_fields = ('play__title',)
    filter_horizontal = ('actors',)
    date_hierarchy = 'date'
    actions = ['export_to_excel', 'delete_selected']

    def export_to_excel(self, request, queryset):
        """Экспорт выбранных представлений в Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Представления"
        
        # Заголовки
        headers = ['Спектакль', 'Режиссер', 'Дата', 'Время', 'Статус', 'Билеты', 'Цена']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Данные
        for row, perf in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=perf.play.title)
            ws.cell(row=row, column=2, value=f"{perf.play.director.first_name} {perf.play.director.last_name}")
            ws.cell(row=row, column=3, value=perf.date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=4, value=perf.date.strftime('%H:%M'))
            ws.cell(row=row, column=5, value=perf.get_status_display())
            ws.cell(row=row, column=6, value=perf.tickets_available)
            ws.cell(row=row, column=7, value=float(perf.ticket_price))
        
        # Стилизация
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = cell.border = ws.cell(row=1, column=1).border
                if cell.row == 1:
                    cell.font = cell.font = ws.cell(row=1, column=1).font
        
        # Автоматическая ширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=performances.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "Экспортировать в Excel"

@admin.register(ActorRole, site=admin_site)
class ActorRoleAdmin(admin.ModelAdmin):
    list_display = ('actor', 'play', 'role_name')
    list_filter = ('play', 'actor')
    search_fields = ('actor__first_name', 'actor__last_name', 'role_name', 'play__title')
    actions = ['delete_selected']
