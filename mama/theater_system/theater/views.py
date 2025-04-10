from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import Group, User
from django.views.decorators.csrf import csrf_protect
from .models import Director, Play, Actor, ActorRole, Casting, Performance, GENDER_CHOICES, GENRE_CHOICES
from .forms import UserRegistrationForm, PlayForm, DirectorForm, ActorForm, CastingForm, PerformanceForm
from django.db.models import Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
from django.db.models import Avg, Count, Case, When, Value, CharField
from django.utils import timezone

def ensure_user_in_director_group(user):
    """Проверяет и добавляет пользователя в группу Director"""
    director_group, created = Group.objects.get_or_create(name='Director')
    if not user.groups.filter(name='Director').exists():
        user.groups.add(director_group)
        user.save()

def is_admin(user):
    return user.is_staff

def is_director(user):
    """Проверка является ли пользователь директором"""
    if user.is_staff:
        return True
    ensure_user_in_director_group(user)
    return user.groups.filter(name='Director').exists()

def is_actor(user):
    return user.groups.filter(name='Actor').exists()

def index(request):
    return render(request, 'theater/index.html')

@csrf_protect
def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Регистрация успешна!')
            return redirect('home')
    else:
        form = UserRegistrationForm()
    return render(request, 'theater/register.html', {'form': form})

@login_required
def home(request):
    return render(request, 'theater/home.html')

@login_required
def director_list(request):
    directors = Director.objects.all()
    
    # Фильтрация по полу
    gender = request.GET.get('gender')
    if gender:
        directors = directors.filter(gender=gender)
    
    # Фильтрация по опыту
    experience = request.GET.get('experience')
    if experience:
        directors = directors.filter(years_of_experience__gte=experience)
    
    # Поиск по имени или фамилии
    search_query = request.GET.get('search')
    if search_query:
        directors = directors.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    context = {
        'directors': directors,
        'selected_gender': gender,
        'selected_experience': experience,
        'search_query': search_query,
    }
    return render(request, 'theater/director_list.html', context)

@login_required
def director_detail(request, pk):
    director = get_object_or_404(Director, pk=pk)
    return render(request, 'theater/director_detail.html', {'director': director})

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff)
def director_create(request):
    if request.method == 'POST':
        form = DirectorForm(request.POST)
        if form.is_valid():
            director = form.save()
            messages.success(request, 'Режиссер успешно добавлен.')
            return redirect('director_list')
    else:
        form = DirectorForm()
    return render(request, 'theater/director_form.html', {'form': form, 'title': 'Добавление режиссера'})

@login_required
def play_list(request):
    plays = Play.objects.all()
    
    # Фильтрация по жанру
    genre = request.GET.get('genre')
    if genre:
        plays = plays.filter(genre=genre)
    
    # Фильтрация по режиссеру
    director_id = request.GET.get('director')
    if director_id:
        plays = plays.filter(director_id=director_id)
    
    # Поиск по названию
    search_query = request.GET.get('search')
    if search_query:
        plays = plays.filter(title__icontains=search_query)
    
    directors = Director.objects.all()
    context = {
        'plays': plays,
        'directors': directors,
        'selected_genre': genre,
        'selected_director': director_id,
        'search_query': search_query,
    }
    return render(request, 'theater/play_list.html', context)

@login_required
def play_detail(request, pk):
    play = get_object_or_404(
        Play.objects.select_related('director').prefetch_related(
            'actorrole_set__actor',
            'performances'
        ), 
        pk=pk
    )
    performances = play.performances.all().order_by('date')
    return render(request, 'theater/play_detail.html', {
        'play': play,
        'performances': performances
    })

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff or is_director(u))
def play_create(request):
    if request.method == 'POST':
        form = PlayForm(request.POST, request.FILES)
        if form.is_valid():
            play = form.save()
            messages.success(request, 'Спектакль успешно добавлен.')
            return redirect('play_detail', pk=play.pk)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = PlayForm()
    return render(request, 'theater/play_form.html', {
        'form': form,
        'title': 'Добавление спектакля',
        'genres': GENRE_CHOICES
    })

@login_required
def actor_list(request):
    actors = Actor.objects.all()
    
    # Фильтрация по полу
    gender = request.GET.get('gender')
    if gender:
        actors = actors.filter(gender=gender)
    
    # Поиск по имени или фамилии
    search_query = request.GET.get('search')
    if search_query:
        actors = actors.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    context = {
        'actors': actors,
        'selected_gender': gender,
        'search_query': search_query,
    }
    return render(request, 'theater/actor_list.html', context)

@login_required
def actor_detail(request, pk):
    actor = get_object_or_404(Actor, pk=pk)
    return render(request, 'theater/actor_detail.html', {'actor': actor})

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff or is_director(u))
def actor_create(request):
    if request.method == 'POST':
        form = ActorForm(request.POST)
        if form.is_valid():
            actor = form.save()
            messages.success(request, 'Актер успешно добавлен.')
            return redirect('actor_list')
    else:
        form = ActorForm()
    return render(request, 'theater/actor_form.html', {'form': form, 'title': 'Добавление актера'})

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff or is_director(u))
def actor_update(request, pk):
    actor = get_object_or_404(Actor, pk=pk)
    if request.method == 'POST':
        form = ActorForm(request.POST, instance=actor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Информация об актере успешно обновлена.')
            return redirect('actor_list')
    else:
        form = ActorForm(instance=actor)
    return render(request, 'theater/actor_form.html', {'form': form, 'actor': actor, 'title': 'Редактирование актера'})

@login_required
def casting_list(request):
    castings = Casting.objects.all()
    
    # Фильтрация по спектаклю
    play_id = request.GET.get('play')
    if play_id:
        castings = castings.filter(play_id=play_id)
    
    # Фильтрация по актеру
    actor_id = request.GET.get('actor')
    if actor_id:
        castings = castings.filter(actor_id=actor_id)
    
    # Фильтрация по статусу
    status = request.GET.get('status')
    if status:
        castings = castings.filter(status=status)
    
    # Фильтрация по дате
    casting_date = request.GET.get('casting_date')
    if casting_date:
        castings = castings.filter(casting_date=casting_date)
    
    plays = Play.objects.all()
    actors = Actor.objects.all()
    context = {
        'castings': castings,
        'plays': plays,
        'actors': actors,
        'selected_play': play_id,
        'selected_actor': actor_id,
        'selected_status': status,
        'selected_date': casting_date,
    }
    return render(request, 'theater/casting_list.html', context)

@login_required
def performance_list(request):
    performances = Performance.objects.all()
    
    # Фильтрация по спектаклю
    play_id = request.GET.get('play')
    if play_id:
        performances = performances.filter(play_id=play_id)
    
    # Фильтрация по дате
    date = request.GET.get('date')
    if date:
        performances = performances.filter(date=date)
    
    # Фильтрация по статусу
    status = request.GET.get('status')
    if status:
        performances = performances.filter(status=status)
    
    plays = Play.objects.all()
    context = {
        'performances': performances,
        'plays': plays,
        'selected_play': play_id,
        'selected_date': date,
        'selected_status': status,
    }
    return render(request, 'theater/performance_list.html', context)

@login_required
def performance_detail(request, pk):
    performance = get_object_or_404(Performance.objects.select_related('play', 'play__director').prefetch_related('actors'), pk=pk)
    return render(request, 'theater/performance_detail.html', {'performance': performance})

@login_required
@csrf_protect
@user_passes_test(is_director)
def performance_update(request, pk):
    """Обновление информации о представлении"""
    performance = get_object_or_404(Performance, pk=pk)
    if request.method == 'POST':
        form = PerformanceForm(request.POST, instance=performance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Информация о представлении успешно обновлена.')
            return redirect('performance_list')
    else:
        form = PerformanceForm(instance=performance)
    return render(request, 'theater/performance_form.html', {
        'form': form,
        'performance': performance,
        'title': 'Редактирование представления'
    })

@login_required
@csrf_protect
@user_passes_test(is_director)
def performance_delete(request, pk):
    """Удаление представления"""
    performance = get_object_or_404(Performance, pk=pk)
    if request.method == 'POST':
        performance.delete()
        messages.success(request, 'Представление успешно удалено.')
        return redirect('performance_list')
    return render(request, 'theater/performance_confirm_delete.html', {'performance': performance})

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff or is_director(u))
def play_update(request, pk):
    play = get_object_or_404(Play, pk=pk)
    if request.method == 'POST':
        form = PlayForm(request.POST, request.FILES, instance=play)
        if form.is_valid():
            form.save()
            messages.success(request, 'Спектакль успешно обновлен.')
            return redirect('play_detail', pk=play.pk)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = PlayForm(instance=play)
    return render(request, 'theater/play_form.html', {
        'form': form,
        'title': 'Редактирование спектакля',
        'play': play,
        'genres': GENRE_CHOICES
    })

@login_required
@csrf_protect
@user_passes_test(is_admin)
def play_delete(request, pk):
    play = get_object_or_404(Play, pk=pk)
    
    if request.method == 'POST':
        play.delete()
        messages.success(request, 'Спектакль успешно удален.')
        return redirect('play_list')
    
    return render(request, 'theater/play_confirm_delete.html', {
        'play': play
    })

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff or is_director(u))
def director_update(request, pk):
    director = get_object_or_404(Director, pk=pk)
    if request.method == 'POST':
        form = DirectorForm(request.POST, instance=director)
        if form.is_valid():
            form.save()
            messages.success(request, 'Информация о режиссере успешно обновлена.')
            return redirect('director_list')
    else:
        form = DirectorForm(instance=director)
    return render(request, 'theater/director_form.html', {'form': form, 'director': director, 'title': 'Редактирование режиссера'})

@login_required
@csrf_protect
@user_passes_test(is_admin)
def director_delete(request, pk):
    director = get_object_or_404(Director, pk=pk)
    
    if request.method == 'POST':
        director.delete()
        messages.success(request, 'Режиссер успешно удален.')
        return redirect('director_list')
    
    return render(request, 'theater/director_confirm_delete.html', {
        'director': director
    })

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff)
def actor_delete(request, pk):
    actor = get_object_or_404(Actor, pk=pk)
    if request.method == 'POST':
        actor.delete()
        messages.success(request, 'Актер успешно удален.')
        return redirect('actor_list')
    return render(request, 'theater/actor_confirm_delete.html', {'actor': actor})

@login_required
@csrf_protect
@user_passes_test(lambda u: u.groups.filter(name='Director').exists())
def casting_update(request, pk):
    casting = get_object_or_404(Casting, pk=pk)
    
    # Проверяем, является ли текущий пользователь режиссером спектакля
    if not request.user.is_staff and casting.play and casting.play.director.user != request.user:
        messages.error(request, 'У вас нет прав для редактирования этого кастинга.')
        return redirect('casting_list')
    
    if request.method == 'POST':
        form = CastingForm(request.POST, instance=casting)
        if form.is_valid():
            form.save()
            messages.success(request, 'Кастинг успешно обновлен.')
            return redirect('casting_list')
    else:
        form = CastingForm(instance=casting)
    return render(request, 'theater/casting_form.html', {'form': form, 'title': 'Редактировать кастинг'})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def export_performances_excel(request):
    """Экспорт списка представлений в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Представления"

    # Определяем стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовки
    headers = [
        'Спектакль',
        'Дата и время',
        'Статус',
        'Цена билета',
        'Доступно билетов',
        'Режиссер'
    ]

    # Применяем стили к заголовкам
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Данные
    performances = Performance.objects.select_related('play', 'play__director').all()
    
    for row, perf in enumerate(performances, 2):
        # Записываем данные
        ws.cell(row=row, column=1, value=perf.play.title)
        ws.cell(row=row, column=2, value=perf.date.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=3, value=dict(Performance.STATUS_CHOICES).get(perf.status, ''))
        ws.cell(row=row, column=4, value=float(perf.ticket_price))
        ws.cell(row=row, column=5, value=perf.tickets_available)
        ws.cell(row=row, column=6, value=f"{perf.play.director.first_name} {perf.play.director.last_name}")

        # Применяем границы к ячейкам
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    # Автоматическая ширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=performances.xlsx'
    
    # Сохраняем файл
    wb.save(response)
    return response

@login_required
@user_passes_test(lambda u: u.is_superuser)
def statistics_view(request):
    # Общая статистика
    total_plays = Play.objects.count()
    total_actors = Actor.objects.count()
    total_performances = Performance.objects.count()
    
    # Статистика по жанрам
    plays_by_genre = Play.objects.values('genre').annotate(count=Count('id'))
    for genre_stat in plays_by_genre:
        if genre_stat['genre'] == 'D':
            genre_stat['genre'] = 'Драма'
        elif genre_stat['genre'] == 'C':
            genre_stat['genre'] = 'Комедия'
        elif genre_stat['genre'] == 'T':
            genre_stat['genre'] = 'Трагедия'
    
    # Статистика по представлениям
    upcoming_performances = Performance.objects.filter(
        date__gt=timezone.now()
    ).count()
    
    avg_ticket_price = Performance.objects.aggregate(
        avg_price=Avg('ticket_price')
    )['avg_price'] or 0
    
    # Статистика по кастингам
    castings_by_status = Casting.objects.values('status').annotate(count=Count('id'))
    for status_stat in castings_by_status:
        if status_stat['status'] == 'pending':
            status_stat['status_display'] = 'На рассмотрении'
        elif status_stat['status'] == 'approved':
            status_stat['status_display'] = 'Утвержден'
        elif status_stat['status'] == 'rejected':
            status_stat['status_display'] = 'Отклонен'
    
    # Популярные спектакли (по количеству представлений)
    popular_plays = Performance.objects.values(
        'play__title'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Самые занятые актеры (по количеству утвержденных кастингов)
    active_actors = Actor.objects.annotate(
        approved_castings_count=Count('castings', filter=Q(castings__status='approved'))
    ).order_by('-approved_castings_count')[:5]
    
    # Статистика за последний месяц
    month_ago = timezone.now() - timedelta(days=30)
    performances_last_month = Performance.objects.filter(
        date__gte=month_ago
    ).count()
    
    new_castings_last_month = Casting.objects.filter(
        casting_date__gte=month_ago
    ).count()
    
    context = {
        'total_plays': total_plays,
        'total_actors': total_actors,
        'total_performances': total_performances,
        'plays_by_genre': plays_by_genre,
        'upcoming_performances': upcoming_performances,
        'avg_ticket_price': avg_ticket_price,
        'castings_by_status': castings_by_status,
        'popular_plays': popular_plays,
        'active_actors': active_actors,
        'performances_last_month': performances_last_month,
        'new_castings_last_month': new_castings_last_month,
    }
    
    return render(request, 'theater/statistics.html', context)

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff or is_director(u))
def performance_create(request):
    if request.method == 'POST':
        form = PerformanceForm(request.POST)
        if form.is_valid():
            performance = form.save()
            messages.success(request, 'Представление успешно добавлено.')
            return redirect('performance_list')
    else:
        form = PerformanceForm()
    return render(request, 'theater/performance_form.html', {'form': form, 'title': 'Добавление представления'})

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff)
def casting_create(request):
    if request.method == 'POST':
        form = CastingForm(request.POST)
        if form.is_valid():
            casting = form.save()
            messages.success(request, 'Кастинг успешно создан.')
            return redirect('casting_detail', pk=casting.pk)
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = CastingForm()
    
    return render(request, 'theater/casting_form.html', {
        'form': form,
        'title': 'Создание кастинга'
    })

@login_required
@csrf_protect
@user_passes_test(lambda u: u.is_staff)
def casting_delete(request, pk):
    casting = get_object_or_404(Casting, pk=pk)
    if request.method == 'POST':
        casting.delete()
        messages.success(request, 'Кастинг успешно удален.')
        return redirect('casting_list')
    return render(request, 'theater/casting_confirm_delete.html', {
        'casting': casting,
        'title': 'Удаление кастинга'
    })

@login_required
@csrf_protect
def casting_detail(request, pk):
    casting = get_object_or_404(Casting, pk=pk)
    return render(request, 'theater/casting_detail.html', {
        'casting': casting,
        'title': f'Кастинг: {casting.role}'
    })

@login_required
@csrf_protect
def casting_confirm_delete(request, pk):
    casting = get_object_or_404(Casting, pk=pk)
    if request.method == 'POST':
        casting.delete()
        messages.success(request, 'Кастинг успешно удален.')
        return redirect('casting_list')
    return render(request, 'theater/casting_confirm_delete.html', {'casting': casting})

# Аналогичные представления для других моделей...
